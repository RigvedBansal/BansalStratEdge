from __future__ import annotations

from datetime import date
from pathlib import Path
from zipfile import ZIP_DEFLATED, ZipFile

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


ROOT = Path(__file__).resolve().parents[1]
DOWNLOAD_DIR = ROOT / "assets" / "toolkit-downloads"
PACK_NAME = "finance-systems-toolkit-pack.zip"

NAVY = "1F3656"
NAVY_SOFT = "304C73"
INK = "243750"
GOLD = "C8A96A"
INPUT_FILL = "FFF2CC"
ACTUAL_FILL = "DDEBF7"
FORMULA_FILL = "E7E6E6"
SUCCESS_FILL = "E2F0D9"
WARNING_FILL = "FCE4D6"
PANEL_FILL = "F7F4EE"
WHITE = "FFFFFF"

THIN_SIDE = Side(style="thin", color="D6DCE5")
THIN_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)
TITLE_FONT = Font(name="Aptos", bold=True, size=18, color=NAVY)
SUBTITLE_FONT = Font(name="Aptos", italic=True, size=10, color=NAVY_SOFT)
HEADER_FONT = Font(name="Aptos", bold=True, size=10, color=WHITE)
LABEL_FONT = Font(name="Aptos", bold=True, size=10, color=NAVY)
BODY_FONT = Font(name="Aptos", size=10, color=INK)
SMALL_FONT = Font(name="Aptos", size=9, color=INK)
HEADER_FILL = PatternFill("solid", fgColor=NAVY)
SUBHEADER_FILL = PatternFill("solid", fgColor="E8EEF6")
INPUT_STYLE = PatternFill("solid", fgColor=INPUT_FILL)
ACTUAL_STYLE = PatternFill("solid", fgColor=ACTUAL_FILL)
FORMULA_STYLE = PatternFill("solid", fgColor=FORMULA_FILL)
SUCCESS_STYLE = PatternFill("solid", fgColor=SUCCESS_FILL)
WARNING_STYLE = PatternFill("solid", fgColor=WARNING_FILL)
PANEL_STYLE = PatternFill("solid", fgColor=PANEL_FILL)
WRAP_TOP = Alignment(vertical="top", wrap_text=True)
CENTER = Alignment(horizontal="center", vertical="center")


def apply_cell_style(cell, *, fill=None, font=None, alignment=None, border=THIN_BORDER, unlocked=False, number_format=None):
    cell.font = font or BODY_FONT
    cell.fill = fill or PatternFill(fill_type=None)
    cell.alignment = alignment or WRAP_TOP
    cell.border = border
    cell.protection = Protection(locked=not unlocked)
    if number_format:
      cell.number_format = number_format


def set_widths(ws, mapping: dict[str, float]) -> None:
    for column, width in mapping.items():
        ws.column_dimensions[column].width = width


def title_block(ws, title: str, subtitle: str, guidance: str) -> None:
    ws["A1"] = title
    apply_cell_style(ws["A1"], font=TITLE_FONT, border=Border())
    ws["A2"] = subtitle
    apply_cell_style(ws["A2"], font=SUBTITLE_FONT, border=Border())
    ws["A3"] = guidance
    apply_cell_style(ws["A3"], fill=PANEL_STYLE, border=THIN_BORDER)
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[3].height = 32


def write_headers(ws, row: int, headers: list[str]) -> None:
    for index, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=index, value=header)
        apply_cell_style(cell, fill=HEADER_FILL, font=HEADER_FONT, alignment=CENTER)


def add_standard_instructions(ws, workbook_title: str, workbook_purpose: str) -> None:
    title_block(
        ws,
        workbook_title,
        "Bansal StratEdge Finance Systems Toolkit workbook",
        workbook_purpose,
    )
    ws["A5"] = "Workbook standards"
    apply_cell_style(ws["A5"], fill=SUBHEADER_FILL, font=LABEL_FONT)
    bullets = [
        "Pale-gold cells are editable assumptions. Blue cells hold sample actuals. Grey cells contain formulas or protected logic.",
        "The workbook follows current planning guidance: operational drivers first, explicit downside/stress scenarios, and short planning cycles that stay decision-ready.",
        "Cash views use direct liquidity lines and weekly cadence where relevant, rather than relying only on P&L logic.",
        "Formula cells are protected and input cells are unlocked so teams can change assumptions without breaking structure.",
        "Adapt the sample data to your context. These are starting points built for speed and clarity.",
    ]
    for offset, bullet in enumerate(bullets, start=6):
        ws[f"A{offset}"] = f"• {bullet}"
        apply_cell_style(ws[f"A{offset}"], border=Border())
    ws["A12"] = "Research cues used in this rebuild"
    apply_cell_style(ws["A12"], fill=SUBHEADER_FILL, font=LABEL_FONT)
    research_notes = [
        "BCG: link strategic, financial, and operational metrics with driver-tree logic and shorter planning cycles.",
        "McKinsey and Deloitte: pair 13-week liquidity views with scenario-based planning rather than treating them as standalone crisis files.",
        "Microsoft Excel guidance: unlock input cells, protect worksheets, and let conditional formatting continue to signal changes.",
    ]
    for offset, note in enumerate(research_notes, start=13):
        ws[f"A{offset}"] = f"• {note}"
        apply_cell_style(ws[f"A{offset}"], border=Border())
    ws.freeze_panes = "A5"
    set_widths(ws, {"A": 120})


def add_score_validation(ws, cell_range: str) -> None:
    dv = DataValidation(type="whole", operator="between", formula1="1", formula2="5", allow_blank=False)
    dv.promptTitle = "Score 1 to 5"
    dv.prompt = "Use 1 for weak / missing discipline and 5 for strong, repeatable discipline."
    dv.errorTitle = "Invalid score"
    dv.error = "Enter an integer score between 1 and 5."
    ws.add_data_validation(dv)
    dv.add(cell_range)


def protect_sheet(ws) -> None:
    ws.protection.sheet = True
    ws.protection.sort = True
    ws.protection.autoFilter = True
    ws.protection.formatCells = False
    ws.protection.formatColumns = False
    ws.protection.formatRows = False
    ws.protection.insertColumns = False
    ws.protection.insertRows = False
    ws.protection.deleteColumns = False
    ws.protection.deleteRows = False
    ws.protection.selectLockedCells = True
    ws.protection.selectUnlockedCells = True


def add_scorecard_conditional_formatting(ws, score_col: str, start_row: int, end_row: int) -> None:
    score_range = f"{score_col}{start_row}:{score_col}{end_row}"
    ws.conditional_formatting.add(
        score_range,
        ColorScaleRule(
            start_type="num",
            start_value=1,
            start_color="F4CCCC",
            mid_type="num",
            mid_value=3,
            mid_color="FFF2CC",
            end_type="num",
            end_value=5,
            end_color="D9EAD3",
        ),
    )


def build_scoring_workbook(filename: str, title: str, purpose: str, sections: list[tuple[str, list[str]]], section_action: str) -> None:
    wb = Workbook()
    instructions = wb.active
    instructions.title = "Instructions"
    add_standard_instructions(instructions, title, purpose)

    scorecard = wb.create_sheet("Scorecard")
    title_block(
        scorecard,
        title,
        "Rate current-state discipline on a 1-5 scale.",
        "Editable inputs are in column C. Section summaries and flags update automatically.",
    )
    write_headers(scorecard, 5, ["Section", "Diagnostic", "Score", "Flag", "Recommended next move"])
    row = 6
    section_ranges = []
    for section_name, questions in sections:
        start_row = row
        for question in questions:
            scorecard.cell(row=row, column=1, value=section_name)
            apply_cell_style(scorecard.cell(row=row, column=1), fill=PANEL_STYLE)
            scorecard.cell(row=row, column=2, value=question)
            apply_cell_style(scorecard.cell(row=row, column=2))
            scorecard.cell(row=row, column=3, value=3)
            apply_cell_style(scorecard.cell(row=row, column=3), fill=INPUT_STYLE, alignment=CENTER, unlocked=True)
            scorecard.cell(row=row, column=4, value=f'=IF(C{row}<=2,"RED",IF(C{row}=3,"AMBER","GREEN"))')
            apply_cell_style(scorecard.cell(row=row, column=4), fill=FORMULA_STYLE, alignment=CENTER)
            scorecard.cell(
                row=row,
                column=5,
                value=f'=IF(C{row}<=2,"{section_action}",IF(C{row}=3,"Tighten ownership and cadence","Keep and document the standard"))',
            )
            apply_cell_style(scorecard.cell(row=row, column=5), fill=FORMULA_STYLE)
            row += 1
        section_ranges.append((section_name, start_row, row - 1))
    add_score_validation(scorecard, f"C6:C{row - 1}")
    add_scorecard_conditional_formatting(scorecard, "C", 6, row - 1)
    set_widths(scorecard, {"A": 26, "B": 60, "C": 11, "D": 14, "E": 34})
    scorecard.freeze_panes = "A6"
    protect_sheet(scorecard)

    summary = wb.create_sheet("Summary")
    title_block(
        summary,
        f"{title} Summary",
        "Executive view",
        "Use this page in leadership review after updating the scorecard.",
    )
    summary["A5"] = "Overall score"
    apply_cell_style(summary["A5"], fill=SUBHEADER_FILL, font=LABEL_FONT)
    summary["B5"] = f"=ROUND(AVERAGE(Scorecard!C6:C{row - 1}),1)"
    apply_cell_style(summary["B5"], fill=FORMULA_STYLE, font=Font(name="Aptos", size=16, bold=True, color=NAVY), alignment=CENTER)
    summary["C5"] = '=IF(B5<2.6,"RED",IF(B5<3.8,"AMBER","GREEN"))'
    apply_cell_style(summary["C5"], fill=FORMULA_STYLE, font=LABEL_FONT, alignment=CENTER)
    write_headers(summary, 8, ["Section", "Average score", "Flag", "Immediate focus"])
    for offset, (section_name, start_row, end_row) in enumerate(section_ranges, start=9):
        summary.cell(offset, 1, section_name)
        apply_cell_style(summary.cell(offset, 1))
        summary.cell(offset, 2, f"=ROUND(AVERAGE(Scorecard!C{start_row}:C{end_row}),1)")
        apply_cell_style(summary.cell(offset, 2), fill=FORMULA_STYLE, alignment=CENTER)
        summary.cell(offset, 3, f'=IF(B{offset}<2.6,"RED",IF(B{offset}<3.8,"AMBER","GREEN"))')
        apply_cell_style(summary.cell(offset, 3), fill=FORMULA_STYLE, alignment=CENTER)
        summary.cell(offset, 4, f'=IF(B{offset}<2.6,"Act in next 30 days",IF(B{offset}<3.8,"Assign clear owner","Maintain cadence"))')
        apply_cell_style(summary.cell(offset, 4), fill=FORMULA_STYLE)
    priority_start = 10 + len(section_ranges)
    summary[f"A{priority_start}"] = "Lowest-scoring priority areas"
    apply_cell_style(summary[f"A{priority_start}"], fill=SUBHEADER_FILL, font=LABEL_FONT)
    for rank in range(1, 4):
        row_index = priority_start + rank
        summary[f"A{row_index}"] = f"Priority {rank}"
        apply_cell_style(summary[f"A{row_index}"], fill=PANEL_STYLE, font=LABEL_FONT)
        summary[f"B{row_index}"] = f'=INDEX($A$9:$A${8 + len(section_ranges)},MATCH(SMALL($B$9:$B${8 + len(section_ranges)},{rank}),$B$9:$B${8 + len(section_ranges)},0))'
        apply_cell_style(summary[f"B{row_index}"], fill=FORMULA_STYLE)
    summary[f"A{priority_start + 5}"] = "Executive summary"
    apply_cell_style(summary[f"A{priority_start + 5}"], fill=SUBHEADER_FILL, font=LABEL_FONT)
    summary[f"A{priority_start + 6}"] = (
        f'=CONCAT("Overall score: ",TEXT(B5,"0.0"),"/5 (",C5,"). Focus first on ",B{priority_start + 1},", ",B{priority_start + 2},'
        f'" and ",B{priority_start + 3},". Use this pack to tighten ownership, refresh cadence, and turn the weakest section into a 30-day action plan.")'
    )
    apply_cell_style(summary[f"A{priority_start + 6}"], fill=PANEL_STYLE)
    set_widths(summary, {"A": 28, "B": 24, "C": 14, "D": 28})
    summary.freeze_panes = "A8"
    protect_sheet(summary)

    output = DOWNLOAD_DIR / filename
    wb.save(output)


def build_rolling_forecast() -> None:
    wb = Workbook()
    instructions = wb.active
    instructions.title = "Instructions"
    add_standard_instructions(
        instructions,
        "Driver-Based Rolling Forecast Starter",
        "A 24-month, driver-based planning pack that links commercial assumptions to revenue, opex, EBITDA, and cash.",
    )

    months = [
        "Apr-26",
        "May-26",
        "Jun-26",
        "Jul-26",
        "Aug-26",
        "Sep-26",
        "Oct-26",
        "Nov-26",
        "Dec-26",
        "Jan-27",
        "Feb-27",
        "Mar-27",
        "Apr-27",
        "May-27",
        "Jun-27",
        "Jul-27",
        "Aug-27",
        "Sep-27",
        "Oct-27",
        "Nov-27",
        "Dec-27",
        "Jan-28",
        "Feb-28",
        "Mar-28",
    ]

    assumptions = wb.create_sheet("Assumptions & Drivers")
    title_block(
        assumptions,
        "Assumptions & Drivers",
        "Change the pale-gold inputs only.",
        "Scenario selector in B5 drives the model. Base values stay visible in column C. Column D shows the scenario-adjusted value used across the pack.",
    )
    assumptions["A5"] = "Scenario"
    apply_cell_style(assumptions["A5"], fill=SUBHEADER_FILL, font=LABEL_FONT)
    assumptions["B5"] = "Base"
    apply_cell_style(assumptions["B5"], fill=INPUT_STYLE, font=LABEL_FONT, alignment=CENTER, unlocked=True)
    scenario_dv = DataValidation(type="list", formula1='"Base,Upside,Downside,Stress"', allow_blank=False)
    assumptions.add_data_validation(scenario_dv)
    scenario_dv.add(assumptions["B5"])
    write_headers(assumptions, 8, ["Driver", "Unit", "Base value", "Scenario-adjusted", "Base", "Upside", "Downside", "Stress"])
    driver_rows = [
        ("New logos per month", "count", 18, 1.0, 1.25, 0.85, 0.65),
        ("ARPA per month", "INR", 48000, 1.0, 1.08, 0.95, 0.9),
        ("Gross retention", "%", 0.93, 1.0, 1.02, 0.98, 0.95),
        ("Expansion rate", "%", 0.04, 1.0, 1.15, 0.85, 0.75),
        ("Services revenue per month", "INR", 1800000, 1.0, 1.15, 0.9, 0.82),
        ("Collections days", "days", 52, 1.0, 0.92, 1.08, 1.2),
        ("Paid CAC", "INR", 42000, 1.0, 0.95, 1.08, 1.18),
        ("INR / USD FX", "rate", 83.4, 1.0, 1.02, 1.05, 1.08),
        ("GST rate", "%", 0.18, 1.0, 1.0, 1.0, 1.0),
        ("Opening ARR", "INR", 96000000, 1.0, 1.0, 1.0, 1.0),
        ("Opening cash", "INR", 74000000, 1.0, 1.0, 1.0, 1.0),
        ("Opening sales headcount", "count", 12, 1.0, 1.0, 1.0, 1.0),
        ("Monthly sales hires", "count", 0.7, 1.0, 1.2, 0.85, 0.65),
        ("Sales cost per head / month", "INR", 280000, 1.0, 1.0, 1.0, 1.0),
        ("Monthly product & eng payroll", "INR", 3200000, 1.0, 1.05, 0.98, 0.95),
        ("Monthly G&A payroll", "INR", 1350000, 1.0, 1.02, 0.98, 0.95),
        ("Cloud spend (USD)", "USD", 18000, 1.0, 1.08, 1.0, 0.95),
        ("Software spend", "INR", 460000, 1.0, 1.02, 0.98, 0.95),
        ("Other opex", "INR", 780000, 1.0, 1.08, 0.94, 0.88),
        ("Monthly capex", "INR", 260000, 1.0, 1.1, 1.0, 0.9),
    ]
    for row_index, (name, unit, base_value, base_mult, upside, downside, stress) in enumerate(driver_rows, start=9):
        assumptions.cell(row_index, 1, name)
        apply_cell_style(assumptions.cell(row_index, 1))
        assumptions.cell(row_index, 2, unit)
        apply_cell_style(assumptions.cell(row_index, 2), fill=PANEL_STYLE)
        assumptions.cell(row_index, 3, base_value)
        apply_cell_style(
            assumptions.cell(row_index, 3),
            fill=INPUT_STYLE,
            alignment=CENTER,
            unlocked=True,
            number_format='0.0%' if unit == "%" else '#,##0.00',
        )
        assumptions.cell(row_index, 4, f'=C{row_index}*INDEX(E{row_index}:H{row_index},1,MATCH($B$5,$E$8:$H$8,0))')
        apply_cell_style(
            assumptions.cell(row_index, 4),
            fill=FORMULA_STYLE,
            alignment=CENTER,
            number_format='0.0%' if unit == "%" else '#,##0.00',
        )
        for column_index, value in enumerate([base_mult, upside, downside, stress], start=5):
            assumptions.cell(row_index, column_index, value)
            apply_cell_style(assumptions.cell(row_index, column_index), fill=INPUT_STYLE, alignment=CENTER, unlocked=True, number_format="0.00x")
    set_widths(assumptions, {"A": 30, "B": 14, "C": 14, "D": 16, "E": 10, "F": 10, "G": 10, "H": 10})
    assumptions.freeze_panes = "A8"
    protect_sheet(assumptions)

    revenue = wb.create_sheet("Revenue Build")
    title_block(
        revenue,
        "Revenue Build",
        "Driver-led revenue mechanics",
        "This sheet links opening ARR, new logos, ARPA, retention, and expansion into monthly revenue and cash collections.",
    )
    revenue["A5"] = "Line item"
    apply_cell_style(revenue["A5"], fill=HEADER_FILL, font=HEADER_FONT, alignment=CENTER)
    for column_index, month in enumerate(months, start=2):
        cell = revenue.cell(row=5, column=column_index, value=month)
        apply_cell_style(cell, fill=HEADER_FILL, font=HEADER_FONT, alignment=CENTER)
    rows = {
        6: "Opening ARR",
        7: "New ARR",
        8: "Churn ARR",
        9: "Expansion ARR",
        10: "Closing ARR",
        11: "SaaS revenue",
        12: "Services revenue",
        13: "Total revenue",
        14: "Cash collections",
    }
    for row_index, label in rows.items():
        revenue.cell(row=row_index, column=1, value=label)
        apply_cell_style(revenue.cell(row=row_index, column=1), fill=PANEL_STYLE)
    for column_index in range(2, 26):
        column_letter = get_column_letter(column_index)
        prev_column = get_column_letter(column_index - 1)
        if column_index == 2:
            revenue[f"{column_letter}6"] = "=Assumptions & Drivers!D18"
        else:
            revenue[f"{column_letter}6"] = f"={prev_column}10"
        revenue[f"{column_letter}7"] = "=Assumptions & Drivers!D9*Assumptions & Drivers!D10"
        revenue[f"{column_letter}8"] = f"={column_letter}6*(1-Assumptions & Drivers!D11)"
        revenue[f"{column_letter}9"] = f"={column_letter}6*Assumptions & Drivers!D12"
        revenue[f"{column_letter}10"] = f"={column_letter}6+{column_letter}7-{column_letter}8+{column_letter}9"
        revenue[f"{column_letter}11"] = f"=({column_letter}6+{column_letter}10)/24"
        revenue[f"{column_letter}12"] = "=Assumptions & Drivers!D13"
        revenue[f"{column_letter}13"] = f"={column_letter}11+{column_letter}12"
        revenue[f"{column_letter}14"] = f"={column_letter}13*(30/(30+Assumptions & Drivers!D14))"
        for row_index in range(6, 15):
            apply_cell_style(revenue[f"{column_letter}{row_index}"], fill=FORMULA_STYLE, number_format="#,##0")
    set_widths(revenue, {"A": 24, **{get_column_letter(i): 11 for i in range(2, 26)}})
    revenue.freeze_panes = "B6"
    protect_sheet(revenue)

    expenses = wb.create_sheet("Expense Build")
    title_block(
        expenses,
        "Expense Build",
        "Keep the cost logic simple enough to challenge.",
        "This version links headcount, CAC, FX, and fixed spend into a monthly opex view.",
    )
    expenses["A5"] = "Line item"
    apply_cell_style(expenses["A5"], fill=HEADER_FILL, font=HEADER_FONT, alignment=CENTER)
    for column_index, month in enumerate(months, start=2):
        cell = expenses.cell(row=5, column=column_index, value=month)
        apply_cell_style(cell, fill=HEADER_FILL, font=HEADER_FONT, alignment=CENTER)
    expense_rows = {
        6: "Sales headcount",
        7: "Sales payroll",
        8: "Product & eng payroll",
        9: "G&A payroll",
        10: "Marketing spend",
        11: "Cloud spend",
        12: "Software spend",
        13: "Other opex",
        14: "Total opex",
        15: "Capex",
    }
    for row_index, label in expense_rows.items():
        expenses.cell(row=row_index, column=1, value=label)
        apply_cell_style(expenses.cell(row=row_index, column=1), fill=PANEL_STYLE)
    for column_index in range(2, 26):
        column_letter = get_column_letter(column_index)
        prev_column = get_column_letter(column_index - 1)
        if column_index == 2:
            expenses[f"{column_letter}6"] = "=Assumptions & Drivers!D20"
        else:
            expenses[f"{column_letter}6"] = f"={prev_column}6+Assumptions & Drivers!D21"
        expenses[f"{column_letter}7"] = f"={column_letter}6*Assumptions & Drivers!D22"
        expenses[f"{column_letter}8"] = "=Assumptions & Drivers!D23"
        expenses[f"{column_letter}9"] = "=Assumptions & Drivers!D24"
        expenses[f"{column_letter}10"] = "=Assumptions & Drivers!D15*Assumptions & Drivers!D9"
        expenses[f"{column_letter}11"] = "=Assumptions & Drivers!D25*Assumptions & Drivers!D16"
        expenses[f"{column_letter}12"] = "=Assumptions & Drivers!D26"
        expenses[f"{column_letter}13"] = "=Assumptions & Drivers!D27"
        expenses[f"{column_letter}14"] = f"=SUM({column_letter}7:{column_letter}13)"
        expenses[f"{column_letter}15"] = "=Assumptions & Drivers!D28"
        for row_index in range(6, 16):
            number_format = "#,##0" if row_index != 6 else "0.0"
            apply_cell_style(expenses[f"{column_letter}{row_index}"], fill=FORMULA_STYLE, number_format=number_format)
    set_widths(expenses, {"A": 24, **{get_column_letter(i): 11 for i in range(2, 26)}})
    expenses.freeze_panes = "B6"
    protect_sheet(expenses)

    outputs = wb.create_sheet("3-Statement Output")
    title_block(
        outputs,
        "3-Statement Output",
        "Management view",
        "This is intentionally concise: enough to steer, not enough to bury the decision in detail.",
    )
    outputs["A5"] = "Line item"
    apply_cell_style(outputs["A5"], fill=HEADER_FILL, font=HEADER_FONT, alignment=CENTER)
    for column_index, month in enumerate(months, start=2):
        cell = outputs.cell(row=5, column=column_index, value=month)
        apply_cell_style(cell, fill=HEADER_FILL, font=HEADER_FONT, alignment=CENTER)
    output_rows = {
        6: "Revenue",
        7: "Gross profit",
        8: "Gross margin %",
        9: "Total opex",
        10: "EBITDA",
        11: "Capex",
        12: "Closing cash",
    }
    for row_index, label in output_rows.items():
        outputs.cell(row=row_index, column=1, value=label)
        apply_cell_style(outputs.cell(row=row_index, column=1), fill=PANEL_STYLE)
    for column_index in range(2, 26):
        column_letter = get_column_letter(column_index)
        prev_column = get_column_letter(column_index - 1)
        outputs[f"{column_letter}6"] = f"='Revenue Build'!{column_letter}13"
        outputs[f"{column_letter}7"] = f"={column_letter}6*0.70"
        outputs[f"{column_letter}8"] = f"=IF({column_letter}6=0,0,{column_letter}7/{column_letter}6)"
        outputs[f"{column_letter}9"] = f"='Expense Build'!{column_letter}14"
        outputs[f"{column_letter}10"] = f"={column_letter}7-{column_letter}9"
        outputs[f"{column_letter}11"] = f"='Expense Build'!{column_letter}15"
        if column_index == 2:
            outputs[f"{column_letter}12"] = f"='Assumptions & Drivers'!D19+{column_letter}10-{column_letter}11"
        else:
            outputs[f"{column_letter}12"] = f"={prev_column}12+{column_letter}10-{column_letter}11"
        for row_index in range(6, 13):
            number_format = "0.0%" if row_index == 8 else "#,##0"
            apply_cell_style(outputs[f"{column_letter}{row_index}"], fill=FORMULA_STYLE, number_format=number_format)
    set_widths(outputs, {"A": 22, **{get_column_letter(i): 11 for i in range(2, 26)}})
    outputs.freeze_panes = "B6"
    protect_sheet(outputs)

    scenarios = wb.create_sheet("Scenarios")
    title_block(
        scenarios,
        "Scenarios",
        "Keep downside and stress alive, not just upside.",
        "The selector in Assumptions & Drivers!B5 switches the pack. Use this page to explain what each scenario assumes.",
    )
    write_headers(scenarios, 5, ["Scenario", "What changes", "How to use it"])
    scenario_rows = [
        ("Base", "Current operating plan with known commitments", "Use in monthly management review"),
        ("Upside", "Faster conversion, better retention, tighter CAC", "Use to test reinvestment choices, not to set default spend"),
        ("Downside", "Slower bookings, longer collections, softer efficiency", "Use before approving discretionary hiring"),
        ("Stress", "Sharp demand shock and tougher collections", "Use for runway, covenant, and contingency planning"),
    ]
    for row_index, values in enumerate(scenario_rows, start=6):
        for column_index, value in enumerate(values, start=1):
            scenarios.cell(row_index, column_index, value)
            apply_cell_style(scenarios.cell(row_index, column_index), fill=PANEL_STYLE if column_index == 1 else None)
    set_widths(scenarios, {"A": 16, "B": 42, "C": 42})
    protect_sheet(scenarios)

    dashboard = wb.create_sheet("Dashboard")
    title_block(
        dashboard,
        "Dashboard",
        "The 6 numbers leadership should debate first.",
        "This page is designed for one monthly slide, not a 20-tab model review.",
    )
    write_headers(dashboard, 5, ["Metric", "Month 1", "Month 12", "Month 24", "Why it matters"])
    dashboard_rows = [
        ("Revenue", "='3-Statement Output'!B6", "='3-Statement Output'!M6", "='3-Statement Output'!Y6", "Tests plan credibility and commercial tempo"),
        ("EBITDA", "='3-Statement Output'!B10", "='3-Statement Output'!M10", "='3-Statement Output'!Y10", "Shows the operating discipline beneath growth"),
        ("Closing cash", "='3-Statement Output'!B12", "='3-Statement Output'!M12", "='3-Statement Output'!Y12", "Anchors runway and financing posture"),
        ("ARR", "='Revenue Build'!B10", "='Revenue Build'!M10", "='Revenue Build'!Y10", "Shows recurring revenue compounding"),
        ("Gross margin %", "='3-Statement Output'!B8", "='3-Statement Output'!M8", "='3-Statement Output'!Y8", "Protects pricing and cloud discipline"),
        ("Sales headcount", "='Expense Build'!B6", "='Expense Build'!M6", "='Expense Build'!Y6", "Links growth ambition to hiring reality"),
    ]
    for row_index, values in enumerate(dashboard_rows, start=6):
        dashboard.cell(row_index, 1, values[0])
        apply_cell_style(dashboard.cell(row_index, 1), fill=PANEL_STYLE)
        for column_index, value in enumerate(values[1:], start=2):
            dashboard.cell(row_index, column_index, value)
            apply_cell_style(
                dashboard.cell(row_index, column_index),
                fill=FORMULA_STYLE if column_index < 5 else None,
                number_format="0.0%" if values[0] == "Gross margin %" and column_index < 5 else "#,##0",
                alignment=CENTER if column_index < 5 else WRAP_TOP,
            )
    set_widths(dashboard, {"A": 22, "B": 14, "C": 14, "D": 14, "E": 38})
    protect_sheet(dashboard)

    wb.save(DOWNLOAD_DIR / "driver-based-rolling-forecast.xlsx")


def build_cash_flow_planner() -> None:
    wb = Workbook()
    instructions = wb.active
    instructions.title = "Instructions"
    add_standard_instructions(
        instructions,
        "13-Week Cash Flow & Runway Planner",
        "A weekly liquidity view with downside logic, covenant awareness, and funding-gap signals.",
    )

    assumptions = wb.create_sheet("Inputs")
    title_block(
        assumptions,
        "Inputs",
        "Weekly liquidity assumptions",
        "Use the scenario selector and the pale-gold cells to refresh cash expectations each week on the same cadence.",
    )
    assumptions["A5"] = "Scenario"
    apply_cell_style(assumptions["A5"], fill=SUBHEADER_FILL, font=LABEL_FONT)
    assumptions["B5"] = "Base"
    apply_cell_style(assumptions["B5"], fill=INPUT_STYLE, font=LABEL_FONT, alignment=CENTER, unlocked=True)
    scenario_dv = DataValidation(type="list", formula1='"Base,Downside,Severe downside"', allow_blank=False)
    assumptions.add_data_validation(scenario_dv)
    scenario_dv.add(assumptions["B5"])
    write_headers(assumptions, 8, ["Driver", "Unit", "Base value", "Scenario-adjusted", "Base", "Downside", "Severe downside"])
    rows = [
        ("Opening cash", "INR", 64000000, 1.0, 1.0, 1.0),
        ("Average weekly collections", "INR", 8200000, 1.0, 0.88, 0.74),
        ("Average weekly payroll", "INR", 3100000, 1.0, 1.0, 1.02),
        ("Average weekly vendors", "INR", 2200000, 1.0, 1.06, 1.12),
        ("Average weekly G&A", "INR", 900000, 1.0, 1.0, 1.04),
        ("Average weekly marketing", "INR", 1250000, 1.0, 0.92, 0.78),
        ("Average weekly capex", "INR", 650000, 1.0, 1.0, 0.82),
        ("Minimum liquidity threshold", "INR", 25000000, 1.0, 1.0, 1.0),
        ("Weekly debt service", "INR", 600000, 1.0, 1.0, 1.0),
    ]
    for row_index, (name, unit, base_value, base_mult, downside, severe) in enumerate(rows, start=9):
        assumptions.cell(row_index, 1, name)
        apply_cell_style(assumptions.cell(row_index, 1))
        assumptions.cell(row_index, 2, unit)
        apply_cell_style(assumptions.cell(row_index, 2), fill=PANEL_STYLE)
        assumptions.cell(row_index, 3, base_value)
        apply_cell_style(assumptions.cell(row_index, 3), fill=INPUT_STYLE, alignment=CENTER, unlocked=True, number_format="#,##0")
        assumptions.cell(row_index, 4, f'=C{row_index}*INDEX(E{row_index}:G{row_index},1,MATCH($B$5,$E$8:$G$8,0))')
        apply_cell_style(assumptions.cell(row_index, 4), fill=FORMULA_STYLE, alignment=CENTER, number_format="#,##0")
        for column_index, value in enumerate([base_mult, downside, severe], start=5):
            assumptions.cell(row_index, column_index, value)
            apply_cell_style(assumptions.cell(row_index, column_index), fill=INPUT_STYLE, alignment=CENTER, unlocked=True, number_format="0.00x")
    set_widths(assumptions, {"A": 28, "B": 12, "C": 14, "D": 16, "E": 10, "F": 12, "G": 14})
    protect_sheet(assumptions)

    weekly = wb.create_sheet("Weekly Cash")
    title_block(
        weekly,
        "Weekly Cash",
        "Direct cash flow view",
        "This is a direct cash forecast. It is meant to support collections action, vendor timing, and runway calls, not just reporting.",
    )
    weekly["A5"] = "Line item"
    apply_cell_style(weekly["A5"], fill=HEADER_FILL, font=HEADER_FONT, alignment=CENTER)
    weeks = [f"Wk {n}" for n in range(1, 14)]
    for column_index, week in enumerate(weeks, start=2):
        weekly.cell(5, column_index, week)
        apply_cell_style(weekly.cell(5, column_index), fill=HEADER_FILL, font=HEADER_FONT, alignment=CENTER)
    labels = {
        6: "Opening cash",
        7: "Collections",
        8: "Other inflows",
        9: "Payroll",
        10: "Vendor payments",
        11: "G&A outflows",
        12: "Marketing",
        13: "Capex",
        14: "Debt service",
        15: "Net movement",
        16: "Closing cash",
        17: "Alert",
    }
    for row_index, label in labels.items():
        weekly.cell(row_index, 1, label)
        apply_cell_style(weekly.cell(row_index, 1), fill=PANEL_STYLE)
    for column_index in range(2, 15):
        column_letter = get_column_letter(column_index)
        prev_column = get_column_letter(column_index - 1)
        if column_index == 2:
            weekly[f"{column_letter}6"] = "=Inputs!D9"
        else:
            weekly[f"{column_letter}6"] = f"={prev_column}16"
        weekly[f"{column_letter}7"] = "=Inputs!D10"
        weekly[f"{column_letter}8"] = "=IF(COLUMN()=5,1200000,0)"
        weekly[f"{column_letter}9"] = "=-Inputs!D11"
        weekly[f"{column_letter}10"] = "=-Inputs!D12"
        weekly[f"{column_letter}11"] = "=-Inputs!D13"
        weekly[f"{column_letter}12"] = "=-Inputs!D14"
        weekly[f"{column_letter}13"] = "=-Inputs!D15"
        weekly[f"{column_letter}14"] = "=-Inputs!D17"
        weekly[f"{column_letter}15"] = f"=SUM({column_letter}7:{column_letter}14)"
        weekly[f"{column_letter}16"] = f"={column_letter}6+{column_letter}15"
        weekly[f"{column_letter}17"] = f'=IF({column_letter}16<Inputs!D16,"Funding gap alert",IF({column_letter}16<Inputs!D16*1.25,"Watch closely","Normal"))'
        for row_index in range(6, 17):
            apply_cell_style(weekly[f"{column_letter}{row_index}"], fill=FORMULA_STYLE, number_format="#,##0")
        apply_cell_style(weekly[f"{column_letter}17"], fill=FORMULA_STYLE, alignment=CENTER)
    weekly.conditional_formatting.add("B16:N16", CellIsRule(operator="lessThan", formula=["Inputs!D16"], fill=WARNING_STYLE))
    set_widths(weekly, {"A": 22, **{get_column_letter(i): 12 for i in range(2, 15)}})
    weekly.freeze_panes = "B6"
    protect_sheet(weekly)

    dashboard = wb.create_sheet("Dashboard")
    title_block(
        dashboard,
        "Dashboard",
        "Immediate liquidity view",
        "Use this page in the weekly cash call. If the file is only reviewed at month-end, it is too late.",
    )
    write_headers(dashboard, 5, ["Metric", "Value", "Interpretation"])
    dashboard_rows = [
        ("Lowest closing cash", "=MIN('Weekly Cash'!B16:N16)", "The tightest point in the 13-week window"),
        ("Week of lowest cash", '=INDEX("Wk "&ROW($1:$13),MATCH(MIN(\'Weekly Cash\'!B16:N16),\'Weekly Cash\'!B16:N16,0))', "The week that needs the most management attention"),
        ("Threshold", "=Inputs!D16", "Minimum liquidity threshold agreed with leadership"),
        ("Weeks below threshold", '=COUNTIF(\'Weekly Cash\'!B16:N16,"<"&Inputs!D16)', "Non-zero means you need a near-term action package"),
        ("Average weekly burn", '=ABS(AVERAGEIF(\'Weekly Cash\'!B15:N15,"<0"))', "Useful for runway framing and financing conversations"),
    ]
    for row_index, values in enumerate(dashboard_rows, start=6):
        for column_index, value in enumerate(values, start=1):
            dashboard.cell(row_index, column_index, value)
            apply_cell_style(
                dashboard.cell(row_index, column_index),
                fill=PANEL_STYLE if column_index == 1 else FORMULA_STYLE if column_index == 2 else None,
                alignment=CENTER if column_index == 2 else WRAP_TOP,
                number_format="#,##0" if column_index == 2 else None,
            )
    set_widths(dashboard, {"A": 26, "B": 18, "C": 44})
    protect_sheet(dashboard)

    wb.save(DOWNLOAD_DIR / "cash-flow-runway-planner.xlsx")


def build_headcount_model() -> None:
    wb = Workbook()
    instructions = wb.active
    instructions.title = "Instructions"
    add_standard_instructions(
        instructions,
        "Headcount & Capacity Planning Model",
        "A hiring model that links headcount decisions to payroll, productivity ramps, and revenue coverage.",
    )

    assumptions = wb.create_sheet("Assumptions")
    title_block(
        assumptions,
        "Assumptions",
        "Editable headcount inputs",
        "Treat hiring as capital allocation. Update role-level assumptions here before touching output tabs.",
    )
    write_headers(assumptions, 5, ["Function", "Opening HC", "Monthly hires", "Monthly exits", "Monthly salary", "Ramp factor", "Monthly revenue / capacity"])
    roles = [
        ("Sales", 12, 1, 0.2, 280000, 0.65, 4100000),
        ("Customer Success", 8, 0.5, 0.1, 190000, 0.8, 0),
        ("Engineering", 18, 0.8, 0.2, 250000, 0.9, 0),
        ("G&A", 7, 0.2, 0.1, 170000, 0.9, 0),
    ]
    for row_index, values in enumerate(roles, start=6):
        for column_index, value in enumerate(values, start=1):
            assumptions.cell(row_index, column_index, value)
            apply_cell_style(
                assumptions.cell(row_index, column_index),
                fill=PANEL_STYLE if column_index == 1 else INPUT_STYLE,
                unlocked=column_index > 1,
                alignment=CENTER if column_index > 1 else WRAP_TOP,
                number_format="#,##0.0" if column_index in {2, 3, 4, 6} else "#,##0",
            )
    set_widths(assumptions, {"A": 20, "B": 12, "C": 12, "D": 12, "E": 14, "F": 12, "G": 22})
    protect_sheet(assumptions)

    months = ["Apr-26", "May-26", "Jun-26", "Jul-26", "Aug-26", "Sep-26", "Oct-26", "Nov-26", "Dec-26", "Jan-27", "Feb-27", "Mar-27"]
    plan = wb.create_sheet("Headcount Plan")
    title_block(
        plan,
        "Headcount Plan",
        "Monthly ending headcount by function",
        "Use this with leadership before approving a hiring wave. The model deliberately keeps the structure tight.",
    )
    plan["A5"] = "Function"
    apply_cell_style(plan["A5"], fill=HEADER_FILL, font=HEADER_FONT, alignment=CENTER)
    for column_index, month in enumerate(months, start=2):
        plan.cell(5, column_index, month)
        apply_cell_style(plan.cell(5, column_index), fill=HEADER_FILL, font=HEADER_FONT, alignment=CENTER)
    for row_index, role in enumerate(roles, start=6):
        plan.cell(row_index, 1, role[0])
        apply_cell_style(plan.cell(row_index, 1), fill=PANEL_STYLE)
        for column_index in range(2, 14):
            column_letter = get_column_letter(column_index)
            prev_column = get_column_letter(column_index - 1)
            source_row = row_index
            if column_index == 2:
                plan[f"{column_letter}{row_index}"] = f"=Assumptions!B{source_row}"
            else:
                plan[f"{column_letter}{row_index}"] = f"={prev_column}{row_index}+Assumptions!C{source_row}-Assumptions!D{source_row}"
            apply_cell_style(plan[f"{column_letter}{row_index}"], fill=FORMULA_STYLE, alignment=CENTER, number_format="#,##0.0")
    set_widths(plan, {"A": 20, **{get_column_letter(i): 11 for i in range(2, 14)}})
    protect_sheet(plan)

    capacity = wb.create_sheet("Capacity View")
    title_block(
        capacity,
        "Capacity View",
        "Connect hiring to real output.",
        "If the revenue plan assumes conversion without commercial capacity, this page should expose the gap immediately.",
    )
    write_headers(capacity, 5, ["Metric", *months])
    lines = ["Productive sales HC", "Revenue capacity", "Customer success load", "Total payroll"]
    for row_index, label in enumerate(lines, start=6):
        capacity.cell(row_index, 1, label)
        apply_cell_style(capacity.cell(row_index, 1), fill=PANEL_STYLE)
    for column_index in range(2, 14):
        column_letter = get_column_letter(column_index)
        capacity[f"{column_letter}6"] = f"='Headcount Plan'!{column_letter}6*Assumptions!F6"
        capacity[f"{column_letter}7"] = f"={column_letter}6*Assumptions!G6"
        capacity[f"{column_letter}8"] = f"='Headcount Plan'!{column_letter}7/220"
        capacity[f"{column_letter}9"] = (
            f"=('Headcount Plan'!{column_letter}6*Assumptions!E6)+('Headcount Plan'!{column_letter}7*Assumptions!E7)+"
            f"('Headcount Plan'!{column_letter}8*Assumptions!E8)+('Headcount Plan'!{column_letter}9*Assumptions!E9)"
        )
        for row_index in range(6, 10):
            apply_cell_style(capacity[f"{column_letter}{row_index}"], fill=FORMULA_STYLE, number_format="#,##0")
    set_widths(capacity, {"A": 22, **{get_column_letter(i): 11 for i in range(2, 14)}})
    protect_sheet(capacity)

    dashboard = wb.create_sheet("Dashboard")
    title_block(
        dashboard,
        "Dashboard",
        "Decision view",
        "Use the dashboard to judge whether the hiring plan creates capacity faster than it consumes cash.",
    )
    write_headers(dashboard, 5, ["Metric", "Opening", "Year-end", "Comment"])
    dashboard_rows = [
        ("Total headcount", "=SUM(Assumptions!B6:B9)", "=SUM('Headcount Plan'!M6:M9)", "Shows whether the organization is becoming heavier than the operating model can support"),
        ("Revenue capacity", "='Capacity View'!B7", "='Capacity View'!M7", "Commercial capacity should be visible before quota plans are approved"),
        ("Monthly payroll", "='Capacity View'!B9", "='Capacity View'!M9", "Tracks the cash impact of the hiring path"),
        ("CS load", "='Capacity View'!B8", "='Capacity View'!M8", "Use to avoid solving growth by creating churn"),
    ]
    for row_index, values in enumerate(dashboard_rows, start=6):
        for column_index, value in enumerate(values, start=1):
            dashboard.cell(row_index, column_index, value)
            apply_cell_style(
                dashboard.cell(row_index, column_index),
                fill=PANEL_STYLE if column_index == 1 else FORMULA_STYLE if column_index in {2, 3} else None,
                number_format="#,##0",
            )
    set_widths(dashboard, {"A": 22, "B": 14, "C": 14, "D": 44})
    protect_sheet(dashboard)

    wb.save(DOWNLOAD_DIR / "headcount-capacity-planning-model.xlsx")


def build_variance_bridge() -> None:
    wb = Workbook()
    instructions = wb.active
    instructions.title = "Instructions"
    add_standard_instructions(
        instructions,
        "Monthly Variance Bridge Template",
        "A bridge pack that moves finance from 'missed budget' to a clearer driver story and action sequence.",
    )

    inputs = wb.create_sheet("Inputs")
    title_block(
        inputs,
        "Inputs",
        "Budget vs actual by movement line",
        "Keep the bridge on one page. Add only the movements that matter enough to change a decision or narrative.",
    )
    write_headers(inputs, 5, ["Movement line", "Budget", "Actual", "Variance", "Commentary cue"])
    movements = [
        ("Budget EBITDA", 7800000, 7800000, "=C6-B6", "Starting point"),
        ("Volume / bookings", 0, -2100000, "=C7-B7", "Pipeline conversion lagged plan"),
        ("Pricing / mix", 0, 640000, "=C8-B8", "Higher enterprise mix partially offset miss"),
        ("Gross margin", 0, -480000, "=C9-B9", "Cloud and onboarding cost pressure"),
        ("Payroll", 0, -720000, "=C10-B10", "Hiring landed before revenue ramp"),
        ("Marketing", 0, 260000, "=C11-B11", "Discretionary spend was cut"),
        ("Actual EBITDA", 5400000, 5400000, "=C12-B12", "Closing point"),
    ]
    for row_index, values in enumerate(movements, start=6):
        for column_index, value in enumerate(values, start=1):
            inputs.cell(row_index, column_index, value)
            apply_cell_style(
                inputs.cell(row_index, column_index),
                fill=PANEL_STYLE if column_index == 1 else INPUT_STYLE if column_index in {2, 3} else FORMULA_STYLE if column_index == 4 else None,
                unlocked=column_index in {2, 3},
                number_format="#,##0",
            )
    protect_sheet(inputs)
    set_widths(inputs, {"A": 24, "B": 14, "C": 14, "D": 14, "E": 42})

    bridge = wb.create_sheet("Bridge")
    title_block(
        bridge,
        "Bridge",
        "Bridge-ready output table",
        "This structure is chart-friendly and commentary-friendly. Use it for board slides or monthly operating reviews.",
    )
    write_headers(bridge, 5, ["Movement line", "Bridge value", "Direction", "Narrative use"])
    for row_index in range(6, 13):
        bridge[f"A{row_index}"] = f"=Inputs!A{row_index}"
        bridge[f"B{row_index}"] = f"=Inputs!D{row_index}"
        bridge[f"C{row_index}"] = f'=IF(B{row_index}<0,"Negative","Positive / neutral")'
        bridge[f"D{row_index}"] = f"=Inputs!E{row_index}"
        for cell_ref in [f"A{row_index}", f"B{row_index}", f"C{row_index}", f"D{row_index}"]:
            apply_cell_style(
                bridge[cell_ref],
                fill=PANEL_STYLE if cell_ref.startswith("A") else FORMULA_STYLE if not cell_ref.startswith("D") else None,
                number_format="#,##0" if cell_ref.startswith("B") else None,
            )
    set_widths(bridge, {"A": 24, "B": 14, "C": 16, "D": 42})
    protect_sheet(bridge)

    commentary = wb.create_sheet("Commentary")
    title_block(
        commentary,
        "Commentary",
        "Draft narrative blocks",
        "BCG's reporting guidance pushes finance from descriptive hindsight into explanatory insight. Use this sheet to keep the commentary short and causal.",
    )
    write_headers(commentary, 5, ["Section", "Draft"])
    commentary["A6"] = "Headline"
    commentary["B6"] = '=CONCAT("EBITDA landed at INR ",TEXT(Inputs!C12/10000000,"0.0")," Cr versus budget at INR ",TEXT(Inputs!B12/10000000,"0.0")," Cr, driven mainly by ",Inputs!A7," and ",Inputs!A10,".")'
    commentary["A7"] = "What changed"
    commentary["B7"] = '=CONCAT("The biggest negative movement came from ",Inputs!A7," while ",Inputs!A8," provided a partial offset.")'
    commentary["A8"] = "What we do next"
    commentary["B8"] = '="Rephase the commercial conversion view, tighten hiring timing against productivity, and keep discretionary spend under weekly review until conversion normalizes."'
    for row_index in range(6, 9):
        apply_cell_style(commentary[f"A{row_index}"], fill=PANEL_STYLE, font=LABEL_FONT)
        apply_cell_style(commentary[f"B{row_index}"], fill=FORMULA_STYLE)
    set_widths(commentary, {"A": 20, "B": 90})
    protect_sheet(commentary)

    wb.save(DOWNLOAD_DIR / "monthly-variance-bridge.xlsx")


def build_driver_decomposition() -> None:
    wb = Workbook()
    instructions = wb.active
    instructions.title = "Instructions"
    add_standard_instructions(
        instructions,
        "Driver Decomposition Worksheet",
        "A worksheet for moving from a topline miss to a cleaner volume, price, and mix explanation.",
    )

    inputs = wb.create_sheet("Inputs")
    title_block(
        inputs,
        "Inputs",
        "Segment data",
        "Load budget and actual for the slices that change management action. Extra rows do not make the story better.",
    )
    write_headers(inputs, 5, ["Segment", "Budget volume", "Actual volume", "Budget price", "Actual price", "Budget mix %", "Actual mix %"])
    segments = [
        ("SMB SaaS", 180, 160, 48000, 47000, 0.35, 0.31),
        ("Mid-market SaaS", 88, 84, 92000, 97000, 0.29, 0.32),
        ("Enterprise SaaS", 26, 21, 235000, 248000, 0.18, 0.17),
        ("Services", 1, 1, 1800000, 1650000, 0.18, 0.20),
    ]
    for row_index, values in enumerate(segments, start=6):
        for column_index, value in enumerate(values, start=1):
            inputs.cell(row_index, column_index, value)
            apply_cell_style(
                inputs.cell(row_index, column_index),
                fill=PANEL_STYLE if column_index == 1 else INPUT_STYLE,
                unlocked=column_index > 1,
                number_format="0.0%" if column_index in {6, 7} else "#,##0",
                alignment=CENTER if column_index > 1 else WRAP_TOP,
            )
    set_widths(inputs, {"A": 18, "B": 14, "C": 14, "D": 14, "E": 14, "F": 12, "G": 12})
    protect_sheet(inputs)

    output = wb.create_sheet("Decomposition")
    title_block(
        output,
        "Decomposition",
        "Volume, price, and mix effects",
        "Use this to show which drivers mattered most. Do not stop at 'revenue missed budget.'",
    )
    write_headers(output, 5, ["Segment", "Budget revenue", "Actual revenue", "Volume effect", "Price effect", "Mix effect", "Action owner"])
    for row_index in range(6, 10):
        output[f"A{row_index}"] = f"=Inputs!A{row_index}"
        output[f"B{row_index}"] = f"=Inputs!B{row_index}*Inputs!D{row_index}"
        output[f"C{row_index}"] = f"=Inputs!C{row_index}*Inputs!E{row_index}"
        output[f"D{row_index}"] = f"=(Inputs!C{row_index}-Inputs!B{row_index})*Inputs!D{row_index}"
        output[f"E{row_index}"] = f"=Inputs!C{row_index}*(Inputs!E{row_index}-Inputs!D{row_index})"
        output[f"F{row_index}"] = f"=(Inputs!G{row_index}-Inputs!F{row_index})*SUM(C$6:C$9)"
        output[f"G{row_index}"] = "Commercial finance"
        for column in "ABCDEFG":
            apply_cell_style(
                output[f"{column}{row_index}"],
                fill=PANEL_STYLE if column == "A" else FORMULA_STYLE if column != "G" else None,
                number_format="#,##0" if column in "BCDEF" else None,
            )
    set_widths(output, {"A": 18, "B": 16, "C": 16, "D": 16, "E": 16, "F": 16, "G": 18})
    protect_sheet(output)

    actions = wb.create_sheet("Action Log")
    title_block(
        actions,
        "Action Log",
        "Turn decomposition into action",
        "Each major driver should have one clear owner and one next move.",
    )
    write_headers(actions, 5, ["Driver", "Issue observed", "Action", "Owner", "Timing"])
    action_rows = [
        ("Volume", "Enterprise close dates slipped", "Rebuild close confidence with deal-by-deal review", "CRO / FP&A", "This month"),
        ("Price", "Discounting rose in SMB", "Reset floor pricing and approval discipline", "Sales ops", "2 weeks"),
        ("Mix", "Services share increased", "Check if delivery pull-forward is hiding product weakness", "CFO", "This month"),
    ]
    for row_index, values in enumerate(action_rows, start=6):
        for column_index, value in enumerate(values, start=1):
            actions.cell(row_index, column_index, value)
            apply_cell_style(actions.cell(row_index, column_index), fill=PANEL_STYLE if column_index == 1 else None)
    set_widths(actions, {"A": 16, "B": 28, "C": 32, "D": 16, "E": 14})
    protect_sheet(actions)

    wb.save(DOWNLOAD_DIR / "driver-decomposition-worksheet.xlsx")


def build_narrative_builder() -> None:
    wb = Workbook()
    instructions = wb.active
    instructions.title = "Instructions"
    add_standard_instructions(
        instructions,
        "Board-Ready Variance Narrative Builder",
        "A workbook for turning variance data into short, decision-ready commentary that boards and CEOs can actually use.",
    )

    inputs = wb.create_sheet("Inputs")
    title_block(
        inputs,
        "Inputs",
        "Five-line narrative builder",
        "Keep the inputs short. If the leadership team needs three pages of explanation, the analysis is probably not clean enough yet.",
    )
    write_headers(inputs, 5, ["Metric", "Budget", "Actual", "Primary driver", "Secondary driver", "Action", "Owner", "Timing"])
    metrics = [
        ("Revenue", 108000000, 101000000, "Enterprise closes slipped", "Paid CAC increased", "Tighten stage-level close review", "CRO", "30 days"),
        ("Gross margin", 0.66, 0.64, "Cloud spend rose", "Implementation mix shifted", "Reset hosting guardrails", "CTO", "30 days"),
        ("EBITDA", 18500000, 14200000, "Revenue miss", "Hiring landed early", "Freeze non-critical backfills", "CFO", "Immediate"),
    ]
    for row_index, values in enumerate(metrics, start=6):
        for column_index, value in enumerate(values, start=1):
            inputs.cell(row_index, column_index, value)
            apply_cell_style(
                inputs.cell(row_index, column_index),
                fill=PANEL_STYLE if column_index == 1 else INPUT_STYLE if column_index in {2, 3, 4, 5, 6, 7, 8} else None,
                unlocked=column_index > 1,
                number_format="0.0%" if row_index == 7 and column_index in {2, 3} else "#,##0" if column_index in {2, 3} else None,
            )
    set_widths(inputs, {"A": 16, "B": 14, "C": 14, "D": 24, "E": 22, "F": 24, "G": 14, "H": 12})
    protect_sheet(inputs)

    narrative = wb.create_sheet("Narrative")
    title_block(
        narrative,
        "Narrative",
        "Structured commentary blocks",
        "Use this output as a draft, then tighten the language for your board or CEO audience.",
    )
    write_headers(narrative, 5, ["Section", "Draft text"])
    narrative_rows = [
        ("Headline", '=CONCAT("The month closed below plan on the back of ",Inputs!D6," and ",Inputs!D8,", with EBITDA landing at INR ",TEXT(Inputs!C8/10000000,"0.0")," Cr.")'),
        ("What happened", '=CONCAT("Revenue closed at INR ",TEXT(Inputs!C6/10000000,"0.0")," Cr versus budget at INR ",TEXT(Inputs!B6/10000000,"0.0")," Cr. Gross margin moved to ",TEXT(Inputs!C7,"0.0%")," as ",Inputs!D7,".")'),
        ("Why it happened", '=CONCAT("The primary root causes were ",Inputs!D6,", ",Inputs!D7,", and the cost timing effect from ",Inputs!D8,".")'),
        ("What we are doing", '=CONCAT(Inputs!F6,"; ",Inputs!F7,"; and ",Inputs!F8,". Owners: ",Inputs!G6,", ",Inputs!G7,", and ",Inputs!G8,".")'),
        ("Board message", '="We are not changing the strategy off one month, but we are tightening execution, timing, and spend until conversion confidence improves."'),
    ]
    for row_index, values in enumerate(narrative_rows, start=6):
        narrative.cell(row_index, 1, values[0])
        narrative.cell(row_index, 2, values[1])
        apply_cell_style(narrative.cell(row_index, 1), fill=PANEL_STYLE, font=LABEL_FONT)
        apply_cell_style(narrative.cell(row_index, 2), fill=FORMULA_STYLE)
    set_widths(narrative, {"A": 18, "B": 100})
    protect_sheet(narrative)

    wb.save(DOWNLOAD_DIR / "variance-narrative-builder.xlsx")


def build_capital_matrix() -> None:
    wb = Workbook()
    instructions = wb.active
    instructions.title = "Instructions"
    add_standard_instructions(
        instructions,
        "Capital Scoring & Prioritization Matrix",
        "A weighted scoring model for ranking capital requests, pilots, growth bets, and efficiency projects on one page.",
    )

    matrix = wb.create_sheet("Matrix")
    title_block(
        matrix,
        "Matrix",
        "Set weights once for the quarter.",
        "Do not change weights to fit a favored project. If the rules keep moving, the matrix becomes theatre.",
    )
    write_headers(matrix, 5, ["Project", "Strategic fit", "ROI / payback", "Capital efficiency", "Risk", "Execution speed", "Weighted score", "Decision"])
    projects = [
        ("Collections automation", 9, 8, 9, 7, 8),
        ("New fintech sales pod", 8, 7, 6, 6, 7),
        ("AI close assistant pilot", 7, 7, 7, 5, 8),
        ("Warehouse expansion", 5, 4, 3, 6, 4),
    ]
    weights = {"B3": 0.25, "C3": 0.25, "D3": 0.20, "E3": 0.15, "F3": 0.15}
    for cell_ref, value in weights.items():
        matrix[cell_ref] = value
        apply_cell_style(matrix[cell_ref], fill=INPUT_STYLE, unlocked=True, alignment=CENTER, number_format="0%")
    matrix["A3"] = "Weights"
    apply_cell_style(matrix["A3"], fill=PANEL_STYLE, font=LABEL_FONT)
    for row_index, values in enumerate(projects, start=6):
        matrix.cell(row_index, 1, values[0])
        apply_cell_style(matrix.cell(row_index, 1), fill=PANEL_STYLE)
        for column_index, value in enumerate(values[1:], start=2):
            matrix.cell(row_index, column_index, value)
            apply_cell_style(matrix.cell(row_index, column_index), fill=INPUT_STYLE, alignment=CENTER, unlocked=True)
        matrix.cell(row_index, 7, f"=SUMPRODUCT(B{row_index}:F{row_index},$B$3:$F$3)")
        apply_cell_style(matrix.cell(row_index, 7), fill=FORMULA_STYLE, alignment=CENTER, number_format="0.0")
        matrix.cell(row_index, 8, f'=IF(G{row_index}>=8,"Fund now",IF(G{row_index}>=7,"Fund in phases",IF(G{row_index}>=6,"Pilot with controls","Kill list")))')
        apply_cell_style(matrix.cell(row_index, 8), fill=FORMULA_STYLE)
    set_widths(matrix, {"A": 24, "B": 14, "C": 14, "D": 16, "E": 10, "F": 16, "G": 14, "H": 18})
    add_score_validation(matrix, "B6:F9")
    protect_sheet(matrix)

    dashboard = wb.create_sheet("Dashboard")
    title_block(
        dashboard,
        "Dashboard",
        "Ranking summary",
        "Use the dashboard to force the funding conversation into one comparable frame.",
    )
    write_headers(dashboard, 5, ["Project", "Weighted score", "Decision", "Payback note"])
    for row_index in range(6, 10):
        dashboard[f"A{row_index}"] = f"=Matrix!A{row_index}"
        dashboard[f"B{row_index}"] = f"=Matrix!G{row_index}"
        dashboard[f"C{row_index}"] = f"=Matrix!H{row_index}"
        dashboard[f"D{row_index}"] = f'=IF(B{row_index}>=8,"Payback likely under 12 months",IF(B{row_index}>=7,"Needs phased investment discipline","Scrutinize before funding"))'
        for column in "ABCD":
            apply_cell_style(
                dashboard[f"{column}{row_index}"],
                fill=PANEL_STYLE if column == "A" else FORMULA_STYLE if column in {"B", "C"} else None,
                number_format="0.0" if column == "B" else None,
            )
    set_widths(dashboard, {"A": 24, "B": 14, "C": 18, "D": 34})
    protect_sheet(dashboard)

    wb.save(DOWNLOAD_DIR / "capital-scoring-prioritization-matrix.xlsx")


def build_ccc_optimizer() -> None:
    wb = Workbook()
    instructions = wb.active
    instructions.title = "Instructions"
    add_standard_instructions(
        instructions,
        "Working Capital & Cash Conversion Cycle Optimizer",
        "A cash-conversion workbook that shows DSO, DIO, DPO, and cash release potential on one page with action levers.",
    )

    inputs = wb.create_sheet("Inputs")
    title_block(
        inputs,
        "Inputs",
        "Current vs target working capital levers",
        "Use peer-relevant targets. A headline CCC number can hide major opportunities inside DSO, DIO, or DPO.",
    )
    write_headers(inputs, 5, ["Lever", "Current", "Target", "Annual base", "Estimated cash release", "Owner"])
    levers = [
        ("DSO", 74, 58, 360000000, "=(B6-C6)/365*D6", "Finance + sales"),
        ("DIO", 49, 35, 220000000, "=(B7-C7)/365*D7", "Supply chain"),
        ("DPO", 38, 50, 220000000, "=(C8-B8)/365*D8", "Procurement + finance"),
    ]
    for row_index, values in enumerate(levers, start=6):
        for column_index, value in enumerate(values, start=1):
            inputs.cell(row_index, column_index, value)
            apply_cell_style(
                inputs.cell(row_index, column_index),
                fill=PANEL_STYLE if column_index == 1 else INPUT_STYLE if column_index in {2, 3, 4, 6} else FORMULA_STYLE,
                unlocked=column_index in {2, 3, 4, 6},
                number_format="#,##0" if column_index in {2, 3, 4, 5} else None,
            )
    protect_sheet(inputs)
    set_widths(inputs, {"A": 14, "B": 10, "C": 10, "D": 16, "E": 20, "F": 20})

    analysis = wb.create_sheet("Analysis")
    title_block(
        analysis,
        "Analysis",
        "Cash conversion cycle math",
        "Keep the logic visible. Cash conversion improvements stick when the business sees the operational cause, not just the finance target.",
    )
    write_headers(analysis, 5, ["Metric", "Value", "Comment"])
    analysis_rows = [
        ("Current CCC", "=Inputs!B6+Inputs!B7-Inputs!B8", "Current days tied up in operations"),
        ("Target CCC", "=Inputs!C6+Inputs!C7-Inputs!C8", "Target operating discipline"),
        ("CCC improvement", "=B6-B7", "Positive number means cash is released"),
        ("Estimated cash release", "=SUM(Inputs!E6:E8)", "Indicative cash unlocked if targets hold"),
    ]
    for row_index, values in enumerate(analysis_rows, start=6):
        for column_index, value in enumerate(values, start=1):
            analysis.cell(row_index, column_index, value)
            apply_cell_style(
                analysis.cell(row_index, column_index),
                fill=PANEL_STYLE if column_index == 1 else FORMULA_STYLE if column_index == 2 else None,
                number_format="#,##0" if column_index == 2 else None,
            )
    set_widths(analysis, {"A": 20, "B": 18, "C": 42})
    protect_sheet(analysis)

    action = wb.create_sheet("Action Tracker")
    title_block(
        action,
        "Action Tracker",
        "Assign owners to each lever",
        "This is where working capital becomes an operating discipline instead of a finance slogan.",
    )
    write_headers(action, 5, ["Lever", "Action", "Owner", "Cadence", "Success metric"])
    action_rows = [
        ("DSO", "Review top 20 receivables weekly and escalate stalled invoices", "Controller", "Weekly", "Overdues fall below 8% of receivables"),
        ("DIO", "Reset safety stock by item family and aged inventory policy", "Ops head", "Fortnightly", "Inventory days trend below target"),
        ("DPO", "Renegotiate top vendor terms and align payment runs", "Procurement", "Monthly", "Average payable days move toward target"),
    ]
    for row_index, values in enumerate(action_rows, start=6):
        for column_index, value in enumerate(values, start=1):
            action.cell(row_index, column_index, value)
            apply_cell_style(action.cell(row_index, column_index), fill=PANEL_STYLE if column_index == 1 else None)
    set_widths(action, {"A": 12, "B": 40, "C": 18, "D": 14, "E": 30})
    protect_sheet(action)

    wb.save(DOWNLOAD_DIR / "cash-conversion-cycle-optimizer.xlsx")


def build_board_pack() -> None:
    wb = Workbook()
    instructions = wb.active
    instructions.title = "Instructions"
    add_standard_instructions(
        instructions,
        "Board & CEO One-Pagers",
        "A compact workbook with a KPI dashboard layout and a decision memo template for capital requests or forecast resets.",
    )

    dashboard = wb.create_sheet("Dashboard")
    title_block(
        dashboard,
        "One-Page Dashboard",
        "Keep the board on one page before appending backup.",
        "This template is meant to compress the story to metrics, 3-4 insight bullets, and the decision pressure underneath them.",
    )
    write_headers(dashboard, 5, ["Metric", "Current", "Prior month", "Plan", "Insight"])
    metrics = [
        ("Revenue", 101000000, 98000000, 108000000, "Conversion softness in enterprise delayed the close."),
        ("Gross margin %", 0.64, 0.65, 0.66, "Cloud and onboarding mix remain the main pressure points."),
        ("EBITDA", 14200000, 15100000, 18500000, "Hiring timing landed before revenue productivity."),
        ("Closing cash", 54800000, 58600000, 61000000, "Cash remains workable, but runway confidence needs downside refresh."),
    ]
    for row_index, values in enumerate(metrics, start=6):
        for column_index, value in enumerate(values, start=1):
            dashboard.cell(row_index, column_index, value)
            apply_cell_style(
                dashboard.cell(row_index, column_index),
                fill=PANEL_STYLE if column_index == 1 else INPUT_STYLE if column_index in {2, 3, 4} else None,
                unlocked=column_index in {2, 3, 4, 5},
                number_format="0.0%" if values[0] == "Gross margin %" and column_index in {2, 3, 4} else "#,##0" if column_index in {2, 3, 4} else None,
            )
    dashboard["A12"] = "Top insights"
    apply_cell_style(dashboard["A12"], fill=SUBHEADER_FILL, font=LABEL_FONT)
    for row_index, text in enumerate(
        [
            "1. The revenue miss is concentrated in enterprise timing, not broad-based demand destruction.",
            "2. Margin pressure is manageable if cloud and implementation costs are governed faster.",
            "3. Hiring should be rephased to the commercial confidence level, not the original plan.",
            "4. Board discussion should center on pace, cash, and what we are willing to defer.",
        ],
        start=13,
    ):
        dashboard[f"A{row_index}"] = text
        apply_cell_style(dashboard[f"A{row_index}"], border=Border())
    set_widths(dashboard, {"A": 18, "B": 16, "C": 16, "D": 16, "E": 48})
    protect_sheet(dashboard)

    memo = wb.create_sheet("Decision Memo")
    title_block(
        memo,
        "Decision Memo",
        "Use for capital requests or forecast resets.",
        "Force the decision to stay concise: what changed, what we are asking for, why now, downside, and how we will track it.",
    )
    memo_fields = [
        ("Request", "Approve phase 1 of collections automation"),
        ("Owner", "CFO + controller"),
        ("Why now", "Receivables aging has become the fastest cash-release lever."),
        ("Capital required", "INR 55 lakh"),
        ("Expected payback", "7 months"),
        ("Downside risk", "Execution delay if ERP integration slips"),
        ("Decision needed", "Approve phase 1 now; defer phase 2 until 60-day review"),
        ("Success measures", "DSO, collections hit rate, manual work hours removed"),
    ]
    write_headers(memo, 5, ["Section", "Draft"])
    for row_index, values in enumerate(memo_fields, start=6):
        memo.cell(row_index, 1, values[0])
        memo.cell(row_index, 2, values[1])
        apply_cell_style(memo.cell(row_index, 1), fill=PANEL_STYLE, font=LABEL_FONT)
        apply_cell_style(memo.cell(row_index, 2), fill=INPUT_STYLE, unlocked=True)
    set_widths(memo, {"A": 18, "B": 86})
    protect_sheet(memo)

    wb.save(DOWNLOAD_DIR / "board-ceo-one-pagers.xlsx")


def build_prompt_library() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Prompt Library"
    title_block(
        ws,
        "AI Prompts for Finance",
        "Copy-paste workbook version",
        "Each row follows Role, Context, Task, Format. Keep a named reviewer on the workflow before any output reaches management or the board.",
    )
    write_headers(ws, 5, ["Category", "Prompt title", "When to use", "Prompt", "Pro tip"])
    prompts = [
        ("Forecasting", "Downside scenario rebuild", "When bookings or collections soften mid-quarter", "Role: CFO advisor. Context: A Series B SaaS company needs a fresh downside view. Task: Rebuild the next 6 months using new logos, retention, CAC, FX, and hiring timing. Format: Table of assumptions, monthly outputs, and top 5 management actions.", "Always ask the model to show which assumptions changed and why."),
        ("Forecasting", "Scenario bridge", "When leadership argues about upside without pressure-testing downside", "Role: FP&A lead. Context: Base, upside, downside, and stress cases already exist. Task: Compare them and identify the 5 assumptions that explain most of the spread. Format: Ranked bridge plus short decision memo.", "Make the model rank the drivers by effect size."),
        ("Forecasting", "Forecast reset memo", "When the old budget is no longer decision-useful", "Role: CFO chief of staff. Context: Actuals have materially diverged from plan. Task: Draft a forecast reset note for leadership covering what changed, what remains intact, and the operating implications. Format: 6 bullets and one paragraph.", "Ask for both 'what changed' and 'what we are not changing'."),
        ("Variance", "Monthly variance commentary", "When actuals close and the team needs a board-ready draft", "Role: Finance business partner. Context: Provide budget, actual, and the top drivers. Task: Draft concise commentary covering what happened, why, and what we do next. Format: Headline, explanation, action.", "Tell the model to keep commentary causal, not descriptive."),
        ("Variance", "Driver decomposition", "When a topline miss is still being discussed too broadly", "Role: FP&A analyst. Context: Segment-level volume, price, and mix data are available. Task: Quantify the main volume, price, mix, and timing drivers. Format: Table plus action note.", "Give the model segment-level data, not just total company totals."),
        ("Variance", "Board narrative builder", "When the board pack needs cleaner language", "Role: CFO advisor. Context: The board wants a short, non-defensive explanation of performance. Task: Draft a calm, precise board narrative with action ownership. Format: 120 words max.", "Ask it to remove adjectives and keep numbers explicit."),
        ("Capital", "Capital request stress test", "Before approving a growth or efficiency project", "Role: Capital committee reviewer. Context: A project deck claims high ROI and quick payback. Task: Identify hidden assumptions, downside risk, and gating criteria. Format: Red flags, questions, recommendation.", "Make the model separate reversible risks from irreversible ones."),
        ("Capital", "Prioritization matrix scorer", "When multiple projects compete for limited cash", "Role: CFO operating partner. Context: Projects are scored on fit, ROI, efficiency, risk, and speed. Task: Rank them and explain the trade-offs. Format: Ranked table and short rationale.", "Require explicit kill-list logic for weak proposals."),
        ("Capital", "Working capital lever finder", "When EBITDA looks fine but cash does not", "Role: Working-capital specialist. Context: DSO, DPO, DIO, aging, terms, and billing patterns are available. Task: Identify the levers most likely to free cash in 30-90 days. Format: Lever, impact, owner, timing.", "Ask for quick wins and structural fixes separately."),
        ("Narrative", "CEO one-pager draft", "Before the monthly CEO review", "Role: CFO chief of staff. Context: Provide 6 KPIs, major variances, and actions. Task: Draft a one-page CEO note. Format: KPI box, 3 insights, 3 actions.", "Limit the note to one screen."),
        ("Narrative", "Decision memo draft", "For investment requests or forecast resets", "Role: Strategy-finance partner. Context: A decision is needed on capital, spend, or plan reset. Task: Draft a decision memo with recommendation, alternatives, downside, and owner. Format: 5 sections.", "Tell the model to state what happens if leadership says no."),
        ("Narrative", "Board Q&A prep", "Before a board meeting", "Role: CFO advisor. Context: The board deck is drafted. Task: Generate likely board questions with crisp answers and backup asks. Format: Question / answer / evidence.", "Ask for hostile but fair questions."),
        ("Workflow", "Close process control design", "When automating recurring finance work", "Role: Finance controls advisor. Context: The team wants to automate part of close or reporting. Task: Design the minimum controls required, including data inputs, reviewer, audit trail, and exceptions. Format: Control matrix.", "Never let AI output skip named human review."),
        ("Workflow", "Prompt QA checklist", "When finance wants repeatable prompt discipline", "Role: Finance transformation lead. Context: A prompt will be reused monthly. Task: Build a QA checklist covering source data, assumptions, checks, and sign-off. Format: Checklist.", "Treat prompt versioning like model versioning."),
        ("Workflow", "Collections prioritization assistant", "When collections teams need better daily triage", "Role: Order-to-cash analyst. Context: Aging and customer-level details are available. Task: Rank invoices for follow-up and suggest the best next touchpoint. Format: Priority table and call notes.", "Keep final outreach human-controlled."),
    ]
    for row_index, values in enumerate(prompts, start=6):
        for column_index, value in enumerate(values, start=1):
            ws.cell(row_index, column_index, value)
            apply_cell_style(ws.cell(row_index, column_index), fill=PANEL_STYLE if column_index == 1 else None)
    set_widths(ws, {"A": 16, "B": 26, "C": 24, "D": 78, "E": 34})
    protect_sheet(ws)
    wb.save(DOWNLOAD_DIR / "finance-ai-prompts-library.xlsx")


def build_simple_support_workbooks() -> None:
    build_scoring_workbook(
        "capital-efficiency-checklist.xlsx",
        "Capital Efficiency Checklist",
        "A 24-line scorecard for cash visibility, working capital, spend discipline, capital allocation, reporting quality, and AI controls.",
        [
            ("Cash visibility and runway", [
                "We refresh a rolling 13-week cash view weekly.",
                "Major near-term cash movements can be explained quickly.",
                "Runway is reviewed alongside downside triggers.",
                "Leadership works from one liquidity truth.",
            ]),
            ("Working capital discipline", [
                "Receivables have named owners and escalation cadence.",
                "Inventory / purchasing follow real demand signals.",
                "Vendor and customer terms are actively managed.",
                "Finance can quantify trapped cash.",
            ]),
            ("Spend control and approvals", [
                "Approval thresholds are current and enforced.",
                "Discretionary spend is reviewed before quarter-end surprises.",
                "Teams understand committed vs optional spend.",
                "Budget owners receive actionable monthly variance views.",
            ]),
            ("Capital allocation quality", [
                "Projects compete on one scoring framework.",
                "Payback and downside are visible before approval.",
                "Underperforming bets are formally re-underwritten.",
                "Cash impact matters as much as headline ROI.",
            ]),
            ("Reporting and decision support", [
                "Leadership reports focus on drivers, not just outputs.",
                "Finance spends more time on analysis than report assembly.",
                "Definitions stay consistent across the company.",
                "Monthly commentary ends with actions and owners.",
            ]),
            ("AI controls in finance", [
                "AI use cases are screened for risk and repeatability.",
                "Source data and prompts are version controlled.",
                "A named reviewer signs off before management use.",
                "Audit trail and exception logging exist.",
            ]),
        ],
        "Assign an owner and fix the process before the next cycle",
    )
    build_scoring_workbook(
        "fpa-system-health-diagnostic.xlsx",
        "FP&A System Health Diagnostic",
        "An 18-question self-audit on model structure, cadence, ownership, and decision usefulness.",
        [
            ("Model structure", [
                "Core models have clear inputs, logic, and outputs.",
                "Assumptions are centralized rather than buried in formulas.",
                "Version control is disciplined and visible.",
                "Driver logic explains plan movement.",
                "Manual overrides are documented.",
            ]),
            ("Cadence & ownership", [
                "Forecast cadence is consistent and leadership-backed.",
                "Each major driver has an owner.",
                "Actuals-to-forecast refresh happens without heroics.",
                "Scenario changes can be run in hours, not weeks.",
                "Review meetings end with decisions, not debate spillover.",
            ]),
            ("Decision usefulness", [
                "Reports explain cause, not just variance.",
                "Cash and P&L views are linked.",
                "Board narratives are concise and action-led.",
                "Leadership trusts the forecast enough to act on it.",
                "Finance is spending more time on analysis than assembly.",
                "The pack supports reallocation decisions quickly.",
                "The file set is usable by more than one person.",
                "Metric definitions are consistent across teams.",
            ]),
        ],
        "Remove friction and tighten the monthly operating rhythm",
    )
    build_scoring_workbook(
        "ai-readiness-finance-audit.xlsx",
        "AI Readiness in Finance Audit",
        "A quick audit that separates real AI opportunity from governance risk inside finance workflows.",
        [
            ("Opportunity fit", [
                "The use case is repetitive enough to benefit from automation.",
                "The output can be reviewed quickly by a named human owner.",
                "The workflow produces enough value to justify change effort.",
                "Inputs are already structured and available.",
            ]),
            ("Data foundation", [
                "Source data is stable and reconciled.",
                "Metric definitions are consistent across systems.",
                "There is a clean handoff into the prompt or automation step.",
                "Exceptions can be traced back to the source.",
            ]),
            ("Governance & control", [
                "A reviewer signs off before management use.",
                "Prompt versions are stored and retrievable.",
                "The team can explain where the output came from.",
                "Sensitive data is appropriately controlled.",
            ]),
            ("Adoption readiness", [
                "The process owner wants the workflow, not just the technology.",
                "The team knows what to do when the model output is wrong.",
                "Fallback manual process still exists.",
                "The use case has a clear success metric.",
            ]),
        ],
        "Pilot only the workflows that are low-regret and reviewable",
    )


def build_zip_pack() -> None:
    pack_path = DOWNLOAD_DIR / PACK_NAME
    with ZipFile(pack_path, "w", ZIP_DEFLATED) as zip_file:
        for workbook_path in sorted(DOWNLOAD_DIR.glob("*.xlsx")):
            zip_file.write(workbook_path, workbook_path.name)
        manifest = "\n".join(workbook.name for workbook in sorted(DOWNLOAD_DIR.glob("*.xlsx")))
        zip_file.writestr(
            "README.txt",
            "Bansal StratEdge Finance Systems Toolkit downloads\n"
            f"Generated: {date.today().isoformat()}\n\n"
            "Files included:\n"
            f"{manifest}\n",
        )


def main() -> None:
    DOWNLOAD_DIR.mkdir(parents=True, exist_ok=True)
    for stale_file in DOWNLOAD_DIR.glob("*"):
        if stale_file.is_file():
            stale_file.unlink()
    build_simple_support_workbooks()
    build_rolling_forecast()
    build_cash_flow_planner()
    build_headcount_model()
    build_variance_bridge()
    build_driver_decomposition()
    build_narrative_builder()
    build_capital_matrix()
    build_ccc_optimizer()
    build_board_pack()
    build_prompt_library()
    build_zip_pack()
    print(f"Generated toolkit downloads in {DOWNLOAD_DIR}")


if __name__ == "__main__":
    main()
